import pandas as pd
import glob2
from tqdm import tqdm
import argparse, os, sys
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import numbers
import re

def get_zip(phone):
    
    phone = str(phone)[0 : 6]
    
    ziplist = pd.read_csv(os.path.join(sys.path[0],'area_to_zip.csv'))
    
    # print('Zip Lookup...')
    found_zip = '#N/A'
    for row in ziplist.itertuples(index=False):
        area_code = row[0]
        zip_code = row[1]
        # print(area_code, zip_code)
        if phone in str(area_code):
            found_zip = str(zip_code)
            # print(found_zip)
            # if '0' in found_zip[0]:
            #     found_zip = '~'+found_zip
            #     print(found_zip)
            break
    return(found_zip)

def get_pwc(title):
    
    pwc_filter = pd.read_csv(os.path.join(sys.path[0],'pwc_filter.csv'))
    
    found_pwc = '0'
    # print('Filtering Pwc ...')
    for row in pwc_filter.itertuples(index=False):
        term = row[0]
        pwc = row[1]
        if term in str(title):
            found_pwc = pwc
            break
    
    return(found_pwc)

def csv_to_upload_sheet(df,output_path):
    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    
    ws.insert_rows(1, amount=4)
    ws.insert_cols(2, amount=1)
    ws.insert_cols(5, amount=6)
    ws['A1'] = 'Prospecting Data'
    ws['B5'] = 'First'
    ws['B4'] = 'Contact Name'
    ws['A4'] = 'Company Name'
    ws['A5'] = '(Required)'
    ws['D4'] = 'Business Phone Number'
    ws['D5'] = '(Required if no Email)'
    ws['E5'] = 'Business Fax'
    ws['F5'] = 'Cell Phone'
    ws['G5'] = 'Website URL'
    ws['H4'] = 'Email'
    ws['H5'] = '(Required if no Business Phone)'
    ws['I5'] = 'Address 1'
    ws['J5'] = 'Address 2'
    ws['M4'] = 'PWC'
    ws['M5'] = '(ID or Description)'
    ws['N5'] = 'Source'
    ws['O5'] = 'Comments'
    for row in ws[2:ws.max_row]: 
        cell = row[11]            # column O
        cell.number_format = numbers.FORMAT_TEXT
        # print(str(cell)[1])
    
    wb.save(output_path + "Filtered_upload_sheet.xlsx")
    
def get_info(path_to_directory,output_path):
    
    dflist = []

    path = path_to_directory + '/*.csv'
    # print (path)
    files = glob2.glob(path)
    
    for file in files:
        # print(file)
        df = pd.read_csv(file)
        try:
            df=df.rename(columns={'title':'company_name','address':'city','Link':'link'})
        except:
            pass
        
        
        dflist.append(df)
    # print(dflist)
    cdf = pd.concat(dflist)
    cdf = cdf.drop_duplicates()
    cdf = cdf.drop_duplicates(subset=['phone'])
    # cdf.to_csv('poop.csv')
    # # print(cdf)
    
    print('getting infos :P')
    
    namelist = []
    phonelist = []
    linklist = []
    citylist = []
    ziplist = []
    pwclist = []
    
    count = 0
    filtered = 0
    good = 0    
    ccdf = ''
    for row in tqdm(cdf.itertuples(index=False), total=cdf.shape[0]):
        # print(row)
        count+=1

        
        name = row[0]
        # print(name)
        # cleaned_name = [character for character in name if character.isalnum()]
        cleaned_name = re.sub(r'\W+', ' ', str(name))
        name = cleaned_name
        # cleaned_name = ''.join(cleaned_name)
        # print(cleaned_name)
        phone = row[1]
        phone = str(phone).strip()
        phone = phone.replace('(','')
        phone = phone.replace(')','')
        phone = phone.replace('-','')
        phone = phone.replace(' ','')
        # print(phone)
        if len(str(phone)) == 10:
            pass
        else:
            filtered+=1
            continue
        try:
            city = row[2]
        except:
            city = ''
        try:
            link = row[3]
        except:
            link = ''
        zipcode = get_zip(phone)
        if zipcode == '#N/A':
            filtered +=1
            continue
        pwc = get_pwc(name)
        if pwc == '0':
            filtered+=1
            continue
        # print(count,'/',len(cdf))
        good+=1
        # print(name, zipcode, pwc)
        namelist.append(name)
        phonelist.append(phone)
        linklist.append(link)
        citylist.append(city)
        ziplist.append(zipcode)
        pwclist.append(pwc)
        cleaned_df = pd.DataFrame({'Company Name':namelist,'Last':linklist, 'Business Phone':phonelist, 'City':citylist,'Zip':ziplist,'PWC':pwclist})
        # cleaned_df = cleaned_df.astype(str)
        ccdf = cleaned_df.drop_duplicates()
    # cleaned_df.to_csv(output_path + 'FilteredList.csv', index=False)
    csv_to_upload_sheet(ccdf, output_path)
    print('Cleaning successful!\n',good, 'good links', filtered, 'bad links filtered out')

    
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--input', dest='input_path', help='specify directory of csv files you wish to filter', action='store', required=True)
    parser.add_argument('-o', '--output', dest='output_path', help='specify where you want output to be saved. default is in current working directory', action='store')
    
    args = parser.parse_args()

    input_directory = args.input_path
    if args.output_path:
        output=args.output_path + '/'
    else:
        output=''
    get_info(input_directory,output)

if __name__ == '__main__':
    main()