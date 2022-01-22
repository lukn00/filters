from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

def csv_to_upload_sheet():
    df = pd.read_csv('FilteredList.csv')
    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    
    ws.insert_rows(1, amount=4)
    ws.insert_cols(2, amount=1)
    ws.insert_cols(5, amount=6)
    ws['A1'] = 'Prospecting Data'
    ws['B5'] = 'First'
    ws['B4'] = 'Contact Name'
    ws['A4'] = 'Company Name'
    ws['A5'] = '(Required)'
    ws['D4'] = 'Business Phone Number'
    ws['D5'] = '(Required if no Email)'
    ws['E5'] = 'Business Fax'
    ws['F5'] = 'Cell Phone'
    ws['G5'] = 'Website URL'
    ws['H4'] = 'Email'
    ws['H5'] = '(Required if no Business Phone)'
    ws['I5'] = 'Address 1'
    ws['J5'] = 'Address 2'
    ws['M4'] = 'PWC'
    ws['M5'] = '(ID or Description)'
    ws['N5'] = 'Source'
    ws['O5'] = 'Comments'
    
    wb.save("pandas_openpyxl.xlsx")
    
if __name__ == "__main__":
    csv_to_upload_sheet()